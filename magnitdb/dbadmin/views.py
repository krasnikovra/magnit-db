from locale import LC_ALL
from math import perm
from django.core.paginator import Paginator
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.template import loader
from django.urls import reverse
from django.contrib.auth import authenticate
from django.contrib.auth import login as djangologin
from django.contrib.auth import logout as djangologout
from django.contrib.auth.models import Permission
from django.contrib.auth.decorators import login_required, permission_required
from .models import *
from openpyxl import Workbook
from datetime import datetime

ROWS_A_PAGE_DOWNLOAD = 2
ROWS_A_PAGE_SEARCH = 2

# all possible filters
filters = [
    'directory_id',
    'department_id',
    'service_id',
    'branch_id',
    'position_id',
]

models = [
    Directory, 
    Department, 
    Service, 
    Branch, 
    Position
]


@login_required
def index(request):
    template = loader.get_template('dbadmin/index.html')
    worker = request.user.worker
    desc_labels = [model._meta.get_field('name').verbose_name for model in models]
    desc_values = [getattr(worker, model.__name__.lower()).name for model in models]
    desc = list(zip(desc_labels, desc_values))

    context = {
        'desc': desc
    }
    return HttpResponse(template.render(context, request))


@login_required
@permission_required('dbadmin.download_data', raise_exception=True)
def download(request):
    # collecting a dict of **kwargs which is something like
    # {
    #   'directory_id__in': request.GET.getlist('directory_id')
    # }

    args = {}
    for filter in filters:
        args[filter + '__in'] = request.GET.getlist(filter)
    # if some GET params are invalid we are aborting any filtering
    workers = Worker.objects.filter(**args).order_by('full_name')
    
    verbose_names = [model._meta.get_field('name').verbose_name for model in models]
    objects = [model.objects.all().order_by('name') for model in models]
    GET = [list(map(int, request.GET.getlist(filter))) for filter in filters]

    paginator = Paginator(workers, ROWS_A_PAGE_DOWNLOAD)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    template = loader.get_template('dbadmin/download.html')
    context = {
        'workers': workers,
        'Models': list(zip(filters, models, objects, verbose_names, GET)),
        'page_obj': page_obj,
    }

    return HttpResponse(template.render(context, request))


@login_required
def export(request):
    # collecting a dict of **kwargs which is something like
    # {
    #   'directory_id__in': request.GET.getlist('directory_id')
    # }
    args = {}
    for filter in filters:
        args[filter + '__in'] = request.GET.getlist(filter)
    workers = Worker.objects.filter(**args).order_by('full_name')

    headers = ('ID', 'ФИО', 'Номер рабочего телефона', 'Номер сотового телефона', 'Дирекция', 'Департамент', 'Служба', 'Отдел', 'Должность')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-dbexport.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    worksheet = workbook.active

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all workers
    for worker in workers:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            worker.pk,
            worker.full_name,
            worker.work_phone,
            worker.cell_phone,
            worker.directory.name,
            worker.department.name,
            worker.service.name,
            worker.branch.name,
            worker.position.name,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def _dispatch_model(model):
    if model == "directory":
        label_text = "Название дирекции"
        verbose_name = "Дирекция" 
        verbose_name_whom = "Дирекцию"
        plural_postfix = "а" #  postfix for the "добавлен" word
        add_perm = "add_directory"
        del_perm = "delete_directory"
        type = Directory
    elif model == "department":
        label_text = "Название департамента"
        verbose_name = "Департамент"
        verbose_name_whom = "Департамент"
        plural_postfix = ""
        add_perm = "add_department"
        del_perm = "delete_department"
        type = Department
    elif model == "service":
        label_text = "Название службы/направления/управления"
        verbose_name = "Служба/Направление/Управление"
        verbose_name_whom = "Службу/Направление/Управление"
        plural_postfix = "а(-о)"
        add_perm = "add_service"
        del_perm = "delete_service"
        type = Service
    elif model == "branch":
        label_text = "Название отдела"
        verbose_name = "Отдел"
        verbose_name_whom = "Отдел"
        plural_postfix = ""
        add_perm = "add_branch"
        del_perm = "delete_branch"
        type = Branch
    elif model == "position":
        label_text = "Название должности"
        verbose_name = "Должность"
        verbose_name_whom = "Должность"
        plural_postfix = "а"
        add_perm = "add_position"
        del_perm = "delete_position"
        type = Position
    else:
        raise Http404()
    return label_text, verbose_name, verbose_name_whom, plural_postfix, add_perm, del_perm, type, 

was_add_succ = False
last_obj_was_add = ""

@login_required
def add(request, model):
    global was_add_succ, last_obj_was_add
    main_view = "add_" + model

    label_text, verbose_name, verbose_name_whom, plural_postfix, add_perm, del_perm, type = _dispatch_model(model)

    if Permission.objects.get(codename=add_perm) not in Permission.objects.filter(group__user=request.user):
        raise PermissionDenied()

    template = loader.get_template('dbadmin/add_generic.html')
    context = {
        'whom': verbose_name_whom,
        'mainview': main_view,
        'model': model,
        'labeltext': label_text,
    }

    if was_add_succ:
        context['msg'] = f"{verbose_name} \"{last_obj_was_add}\" успешно добавлен{plural_postfix}."
        context['err'] = False
        was_add_succ = False

    return HttpResponse(template.render(context, request))


@login_required
def add_save(request, model):
    global was_add_succ, last_obj_was_add
    main_view = "add_" + model

    label_text, verbose_name, verbose_name_whom, plural_postfix, add_perm, del_perm, type = _dispatch_model(model)

    name = request.POST.get('name', None)
    template = loader.get_template('dbadmin/add_generic.html')

    if name is None:
        raise Http404()

    if name == "":
        context = {
            'whom': verbose_name_whom,
            'mainview': main_view,
            'model': model,
            'labeltext': label_text,
            'msg': f"Произошла ошибка: поле \"{label_text}\" пусто.",
            'err': True
        }
        return HttpResponse(template.render(context, request))

    obj = type(name=name)
    obj.save()

    was_add_succ = True
    last_obj_was_add = obj.name

    return HttpResponseRedirect(reverse('dbadmin:add', args=[model]))

was_delete_succ = False
last_obj_was_delete = ""

@login_required
def delete(request, model):
    global was_delete_succ, last_obj_was_add
    main_view = "delete_" + model

    label_text, verbose_name, verbose_name_whom, plural_postfix, add_perm, del_perm, type = _dispatch_model(model)

    if Permission.objects.get(codename=del_perm) not in Permission.objects.filter(group__user=request.user):
        raise PermissionDenied()

    template = loader.get_template('dbadmin/delete_generic.html')

    context = {
        'whom': verbose_name_whom,
        'mainview': main_view,
        'model': model,
        'labeltext': label_text,
        'objects': type.objects.all()
    }

    if was_delete_succ:
        context['msg'] = f"{verbose_name} \"{last_obj_was_delete}\" успешно удален{plural_postfix}."
        context['err'] = False
        was_delete_succ = False

    return HttpResponse(template.render(context, request))

@login_required
def delete_save(request, model):
    global was_delete_succ, last_obj_was_delete
    main_view = "delete_" + model

    label_text, verbose_name, verbose_name_whom, plural_postfix, add_perm, del_perm, type = _dispatch_model(model)

    obj_id = request.POST.get('objid', None)
    template = loader.get_template('dbadmin/delete_generic.html')

    if obj_id is None:
        raise Http404()

    context = {
        'whom': verbose_name_whom,
        'mainview': main_view,
        'model': model,
        'labeltext': label_text,
        'objects': type.objects.all()
    }

    obj = type.objects.get(pk=obj_id)
    kwargs = {}
    kwargs[model + '_id'] = obj_id
    if Worker.objects.filter(**kwargs):
        context['msg'] = 'Существуют работники, для которых ' + verbose_name + " - " + obj.name + "."
        context['err'] = True
        return HttpResponse(template.render(context, request))

    obj.delete()

    was_delete_succ = True
    last_obj_was_delete = obj.name

    return HttpResponseRedirect(reverse('dbadmin:delete', args=[model]))

was_edit_succ = False
last_obj_was_edit_old = ''
last_obj_was_edit_new = ''

@login_required
def edit(request, model):
    global was_edit_succ, last_obj_was_edit_old, last_obj_was_edit_new
    main_view = "edit_" + model

    label_text, verbose_name, verbose_name_whom, plural_postfix, add_perm, del_perm, type = _dispatch_model(model)

    template = loader.get_template('dbadmin/edit_generic.html')
    context = {
        'whom': verbose_name_whom,
        'mainview': main_view,
        'model': model,
        'labeltext': label_text,
        'labeltext_new': "Новое " + label_text.lower(),
        'objects': type.objects.all()
    }

    if was_edit_succ:
        context['msg'] = f"{verbose_name} \"{last_obj_was_edit_old}\" успешно изменен{plural_postfix} на \"{last_obj_was_edit_new}\"."
        context['err'] = False
        was_edit_succ = False

    return HttpResponse(template.render(context, request))

@login_required
def edit_save(request, model):
    global was_edit_succ, last_obj_was_edit_old, last_obj_was_edit_new
    main_view = "edit_" + model

    label_text, verbose_name, verbose_name_whom, plural_postfix, add_perm, del_perm, type = _dispatch_model(model)
    label_text_new = "Новое " + label_text.lower()

    template = loader.get_template('dbadmin/edit_generic.html')
    context = {
        'whom': verbose_name_whom,
        'mainview': main_view,
        'model': model,
        'labeltext': label_text,
        'labeltext_new': "Новое " + label_text.lower(),
        'objects': type.objects.all()
    }

    obj_id = request.POST.get('objid', None)
    new_name = request.POST.get('name', None)
    if obj_id is None or new_name is None:
        raise Http404()
    if new_name == "":
        context['msg'] = f"Произошла ошибка: поле \"{label_text_new}\" пусто."
        context['err'] = True
        return HttpResponse(template.render(context, request))

    obj = type.objects.get(pk=obj_id)
    last_obj_was_edit_old = obj.name
    obj.name = new_name
    obj.save()
    was_edit_succ = True
    last_obj_was_edit_new = obj.name
    
    return HttpResponseRedirect(reverse('dbadmin:edit', args=[model]))

@login_required
@permission_required('dbadmin.search_worker', raise_exception=True)
def profile(request, worker_id):
    try:
        worker = Worker.objects.get(pk=worker_id)
    except Worker.DoesNotExist:
        raise Http404()
    template = loader.get_template('dbadmin/profile.html')
    desc_labels = [model._meta.get_field('name').verbose_name for model in models]
    desc_values = [getattr(worker, model.__name__.lower()).name for model in models]
    desc = list(zip(desc_labels, desc_values))
    context = {
        'worker': worker,
        'desc': desc
    }
    return HttpResponse(template.render(context, request))


@login_required
@permission_required('dbadmin.edit_worker', raise_exception=True)
def profile_edit(request, worker_id):
    try:
        worker = Worker.objects.get(pk=worker_id)
    except Worker.DoesNotExist:
        raise Http404()
    template = loader.get_template('dbadmin/profile_edit.html')
    desc_labels = [model._meta.get_field('name').verbose_name for model in models]
    objects = [model.objects.all() for model in models]
    current_obj_ids = [getattr(worker, filter) for filter in filters]
    data = list(zip(desc_labels, filters, objects, current_obj_ids))
    context = {
        'worker': worker,
        'data': data
    }
    return HttpResponse(template.render(context, request))


@login_required
def profile_edit_save(request, worker_id):
    try:
        worker = Worker.objects.get(pk=worker_id)
    except Worker.DoesNotExist:
        raise Http404()
    template = loader.get_template('dbadmin/profile_edit.html')
    desc_labels = [model._meta.get_field('name').verbose_name for model in models]
    objects = [model.objects.all() for model in models]
    current_obj_ids = [getattr(worker, filter) for filter in filters]
    data = list(zip(desc_labels, filters, objects, current_obj_ids))
    context = {
        'worker': worker,
        'data': data
    }

    full_name = request.POST.get('full_name', None)
    cell_phone = request.POST.get('cell_phone', None)
    work_phone = request.POST.get('work_phone', None)
    directory_id = request.POST.get('directory_id', None)
    department_id = request.POST.get('department_id', None)
    service_id = request.POST.get('service_id', None)
    branch_id = request.POST.get('branch_id', None)
    position_id = request.POST.get('position_id', None)

    if full_name is None or cell_phone is None or work_phone is None or directory_id is None or department_id is None or service_id is None or branch_id is None or position_id is None:
        raise Http404()

    if full_name == '' or cell_phone == '' or work_phone == '':
        context['msg'] = "Произошла ошибка. Одно и полей оказалось пусто."
        context['err'] = True
        return HttpResponse(template.render(context, request))

    worker.full_name = full_name
    worker.cell_phone = cell_phone
    worker.work_phone = work_phone
    worker.directory = Directory.objects.get(pk=directory_id)
    worker.department = Department.objects.get(pk=department_id)
    worker.service = Service.objects.get(pk=service_id)
    worker.branch = Branch.objects.get(pk=branch_id)
    worker.position = Position.objects.get(pk=position_id)
    worker.save()

    return HttpResponseRedirect(reverse('dbadmin:profile', args=[worker_id]))


@login_required
@permission_required('dbadmin.assign_group', raise_exception=True)
def profile_assign_group(request, worker_id):
    try:
        worker = Worker.objects.get(pk=worker_id)
    except Worker.DoesNotExist:
        raise Http404()
    template = loader.get_template('dbadmin/profile_assign_group.html')
    desc_labels = [model._meta.get_field('name').verbose_name for model in models]
    desc_values = [getattr(worker, model.__name__.lower()).name for model in models]
    desc = list(zip(desc_labels, desc_values))
    context = {
        'worker': worker,
        'desc': desc,
        'groups': Group.objects.all()
    }
    return HttpResponse(template.render(context, request))


@login_required
def profile_assign_group_save(request, worker_id):
    try:
        worker = Worker.objects.get(pk=worker_id)
    except Worker.DoesNotExist:
        raise Http404()

    group_ids = request.POST.getlist('group_id', None)
    if group_ids is None:
        raise Http404()
    worker.user.groups.set(Group.objects.filter(pk__in=group_ids))

    return HttpResponseRedirect(reverse('dbadmin:profile', args=[worker_id]))


@login_required
@permission_required('dbadmin.search_worker', raise_exception=True)
def search(request):
    kwargs = {}
    full_name = request.GET.get('full_name', None)
    workers_and_descs = []
    page_obj = None
    collapse_filters = True
    if full_name is not None:
        kwargs['full_name__icontains'] = full_name
        for filter in filters:
            filter_get = request.GET.getlist(filter)
            if filter_get:
                collapse_filters = False
                kwargs[filter + '__in'] = request.GET.getlist(filter)
        workers = Worker.objects.filter(**kwargs).order_by('full_name')
        desc_values = [[getattr(worker, model.__name__.lower()).name for model in models][::-1] for worker in workers]
        workers_and_descs = list(zip(workers, desc_values))

        paginator = Paginator(workers_and_descs, ROWS_A_PAGE_SEARCH)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

    verbose_names = [model._meta.get_field('name').verbose_name for model in models]
    objects = [model.objects.all().order_by('name') for model in models]
    GET = [list(map(int, request.GET.getlist(filter))) for filter in filters]

    template = loader.get_template("dbadmin/search.html")
    context = {
        'collapse_filters': collapse_filters,
        'Models': list(zip(filters, models, objects, verbose_names, GET)),
        'results_count': len(workers_and_descs),
        'page_obj': page_obj,
        'full_name': full_name if full_name is not None else ''
    }
    return HttpResponse(template.render(context, request))

was_group_add_succ = False
last_group_was_add = ''

@login_required
@permission_required('dbadmin.add_group', raise_exception=True)
def group_add(request):
    global was_group_add_succ, last_group_was_add
    main_view = "group_add"

    label_text = "Название группы"
    verbose_name = "Группа" 
    verbose_name_whom = "Группу"
    plural_postfix = "a"

    model_permissions = []
    for model in models + [GroupAuxPerm, Worker]:
        model_permissions += [
        (model._meta.verbose_name,
         Permission.objects.filter(content_type__app_label='dbadmin', content_type__model=model._meta.model_name))
        ]

    template = loader.get_template('dbadmin/group_add.html')
    context = {
        'whom': verbose_name_whom,
        'mainview': main_view,
        'labeltext': label_text,
        'model_permissions': model_permissions,
    }

    if was_group_add_succ:
        context['msg'] = f"{verbose_name} \"{last_group_was_add}\" успешно добавлен{plural_postfix}."
        context['err'] = False
        was_group_add_succ = False

    return HttpResponse(template.render(context, request))

@login_required
def group_add_save(request):
    global was_group_add_succ, last_group_was_add
    main_view = "group_add"

    label_text = "Название группы"
    verbose_name = "Группа" 
    verbose_name_whom = "Группу"
    plural_postfix = "a"

    name = request.POST.get('name', None)
    perm_ids = request.POST.getlist('perm_id', None)
    template = loader.get_template('dbadmin/group_add.html')

    model_permissions = []
    for model in models + [GroupAuxPerm, Worker]:
        model_permissions += [
        (model._meta.verbose_name,
         Permission.objects.filter(content_type__app_label='dbadmin', content_type__model=model._meta.model_name))
        ]

    if name is None or perm_ids is None:
        raise Http404()

    if name == "":
        context = {
            'whom': verbose_name_whom,
            'mainview': main_view,
            'model': model,
            'labeltext': label_text,
            'model_permissions': model_permissions,
            'msg': f"Произошла ошибка: поле \"{label_text}\" пусто.",
            'err': True
        }
        return HttpResponse(template.render(context, request))

    group = Group(name=name)
    group.save()
    group.permissions.set(Permission.objects.filter(pk__in=perm_ids))

    was_group_add_succ = True
    last_group_was_add = group.name

    return HttpResponseRedirect(reverse('dbadmin:group_add'))


was_group_delete_succ = False
last_group_was_delete = ''

@login_required
@permission_required('dbadmin.delete_group', raise_exception=True)
def group_delete(request):
    global was_group_delete_succ, last_group_was_delete

    template = loader.get_template('dbadmin/group_delete.html')

    context = {
        'groups': Group.objects.all()
    }

    if was_group_delete_succ:
        context['msg'] = f"Группа \"{last_group_was_delete}\" успешно удалена."
        context['err'] = False
        was_group_delete_succ = False

    return HttpResponse(template.render(context, request))

@login_required
def group_delete_save(request):
    global was_group_delete_succ, last_group_was_delete

    group_id = request.POST.get('group_id', None)

    if group_id is None:
        raise Http404()

    obj = Group.objects.get(pk=group_id)
    obj.delete()

    was_group_delete_succ = True
    last_group_was_delete = obj.name

    return HttpResponseRedirect(reverse('dbadmin:group_delete'))

was_group_edit_succ = False
last_group_was_edit_old_name = ''
last_group_was_edit_new_name = ''

@login_required
@permission_required('dbadmin.edit_group', raise_exception=True)
def group_edit(request):
    global was_group_edit_succ, last_group_was_edit_old_name, last_group_was_edit_new_name

    model_permissions = []
    for model in models + [GroupAuxPerm, Worker]:
        model_permissions += [
        (model._meta.verbose_name,
         Permission.objects.filter(content_type__app_label='dbadmin', content_type__model=model._meta.model_name))
        ]

    template = loader.get_template('dbadmin/group_edit.html')
    context = {
        'groups': Group.objects.all(),
        'model_permissions': model_permissions,
    }

    if was_group_edit_succ:
        context['msg'] = f"Группа \"{last_group_was_edit_old_name}\" успешно изменена"
        if last_group_was_edit_old_name != last_group_was_edit_new_name:
            context['msg'] += f' на \"{last_group_was_edit_new_name}\"'
        context['msg'] += '.'
        context['err'] = False
        was_group_edit_succ = False

    return HttpResponse(template.render(context, request))

@login_required
def group_edit_save(request):
    global was_group_edit_succ, last_group_was_edit_old_name, last_group_was_edit_new_name

    new_name = request.POST.get('name', None)
    group_id = request.POST.get('group_id', None)
    perm_ids = request.POST.getlist('perm_id', None)
    template = loader.get_template('dbadmin/group_edit.html')

    model_permissions = []
    for model in models + [GroupAuxPerm, Worker]:
        model_permissions += [
        (model._meta.verbose_name,
         Permission.objects.filter(content_type__app_label='dbadmin', content_type__model=model._meta.model_name))
        ]
    if new_name is None or perm_ids is None:
        raise Http404()

    context = {
        'groups': Group.objects.all(),
        'model_permissions': model_permissions,
    }

    if group_id is None:
        context['msg'] = f"Произошла ошибка: изменяемая группа не выбрана."
        context['err'] = True
        return HttpResponse(template.render(context, request))

    if new_name == "":
        context['msg'] = f"Произошла ошибка: поле \"Новое название группы\" пусто."
        context['err'] = True
        return HttpResponse(template.render(context, request))

    group = Group.objects.get(pk=group_id)
    last_group_was_edit_old_name = group.name
    group.name = new_name
    group.permissions.set(Permission.objects.filter(pk__in=perm_ids))
    group.save()

    was_group_edit_succ = True
    last_group_was_edit_new_name = group.name

    return HttpResponseRedirect(reverse('dbadmin:group_edit'))

@login_required
def group(request, group_id):
    try:
        group = Group.objects.get(pk=group_id)
    except Group.DoesNotExist:
        raise Http404()

    template = loader.get_template('dbadmin/group.html')
    context = {
        'group': group
    }
    return HttpResponse(template.render(context, request))

was_cellphone_edit_succ = False
last_cellphone_was_edit = ''

@login_required
def changecellphone(request):
    global was_cellphone_edit_succ, last_cellphone_was_edit
    template = loader.get_template('dbadmin/changecellphone.html')
    context = {}
    if was_cellphone_edit_succ:
        context['msg'] = f'Номер сотового телефона успешно изменен на {last_cellphone_was_edit}.'
        context['err'] = False
        was_cellphone_edit_succ = False
    return HttpResponse(template.render(context, request))
    
@login_required
def changecellphone_save(request):
    global was_cellphone_edit_succ, last_cellphone_was_edit
    new_cell_phone = request.POST.get('cellphone', None)
    if new_cell_phone is None:
        raise Http404()
    if new_cell_phone == '':
        context = {
            'msg': "Произошла ошибка: поле \"Новый номер сотового телефона\" пусто.",
            'err': True
        }
        template = loader.get_template('dbadmin/changecellphone.html')
        return HttpResponse(template.render(context, request))
    worker = request.user.worker
    worker.cell_phone = new_cell_phone
    worker.save()
    was_cellphone_edit_succ = True
    last_cellphone_was_edit = new_cell_phone
    return HttpResponseRedirect(reverse('dbadmin:changecellphone'))


def login(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('dbadmin:index'))

    template = loader.get_template('dbadmin/login.html')
    context = {}
    return HttpResponse(template.render(context, request))
    

def login_confirm(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('dbadmin:index'))

    username = request.POST.get('username', None)
    password = request.POST.get('password', None)
    
    user = authenticate(username=username, password=password)

    if user is not None:
        djangologin(request, user)
        return HttpResponseRedirect(reverse('dbadmin:index'))

    template = loader.get_template('dbadmin/login.html')
    context = {
        'msg': "Неверный логин или пароль."
    }
    return HttpResponse(template.render(context, request))


@login_required
def logout(request):
    djangologout(request)
    return HttpResponseRedirect(reverse('dbadmin:login'))